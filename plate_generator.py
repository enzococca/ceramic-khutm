#!/usr/bin/env python3
"""
Plate Generator Module
Generates archaeological publication plates (PDF A4) with pottery images
organized by period and optionally by stratigraphic unit (US).
"""

import os
import io
import base64
import tempfile
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# A4 dimensions in points (1 point = 1/72 inch)
A4_WIDTH, A4_HEIGHT = A4  # 595.27, 841.89

# Margins in mm
MARGIN_TOP = 25 * mm
MARGIN_BOTTOM = 20 * mm
MARGIN_LEFT = 20 * mm
MARGIN_RIGHT = 20 * mm

# Usable area
USABLE_WIDTH = A4_WIDTH - MARGIN_LEFT - MARGIN_RIGHT
USABLE_HEIGHT = A4_HEIGHT - MARGIN_TOP - MARGIN_BOTTOM

# Caption settings
CAPTION_FONT = 'Helvetica'
CAPTION_SIZE = 9
CAPTION_HEIGHT = 15  # points below image for caption


# Predefined layouts (8 layouts)
# Coordinates are in points from bottom-left corner
# Each cell: {'x': left, 'y': bottom, 'w': width, 'h': height}
PREDEFINED_LAYOUTS = {
    '2x2': {
        'id': '2x2',
        'name': 'Grid 2x2',
        'description': '4 equal images',
        'cells': 4,
        'grid': [
            {'x': MARGIN_LEFT, 'y': A4_HEIGHT - MARGIN_TOP - 360, 'w': USABLE_WIDTH/2 - 5, 'h': 340},
            {'x': MARGIN_LEFT + USABLE_WIDTH/2 + 5, 'y': A4_HEIGHT - MARGIN_TOP - 360, 'w': USABLE_WIDTH/2 - 5, 'h': 340},
            {'x': MARGIN_LEFT, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/2 - 5, 'h': 340},
            {'x': MARGIN_LEFT + USABLE_WIDTH/2 + 5, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/2 - 5, 'h': 340},
        ]
    },
    '2x3': {
        'id': '2x3',
        'name': 'Grid 2x3',
        'description': '6 medium images',
        'cells': 6,
        'grid': [
            {'x': MARGIN_LEFT, 'y': A4_HEIGHT - MARGIN_TOP - 240, 'w': USABLE_WIDTH/2 - 5, 'h': 220},
            {'x': MARGIN_LEFT + USABLE_WIDTH/2 + 5, 'y': A4_HEIGHT - MARGIN_TOP - 240, 'w': USABLE_WIDTH/2 - 5, 'h': 220},
            {'x': MARGIN_LEFT, 'y': A4_HEIGHT - MARGIN_TOP - 480, 'w': USABLE_WIDTH/2 - 5, 'h': 220},
            {'x': MARGIN_LEFT + USABLE_WIDTH/2 + 5, 'y': A4_HEIGHT - MARGIN_TOP - 480, 'w': USABLE_WIDTH/2 - 5, 'h': 220},
            {'x': MARGIN_LEFT, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/2 - 5, 'h': 220},
            {'x': MARGIN_LEFT + USABLE_WIDTH/2 + 5, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/2 - 5, 'h': 220},
        ]
    },
    '3x3': {
        'id': '3x3',
        'name': 'Grid 3x3',
        'description': '9 small images',
        'cells': 9,
        'grid': [
            # Row 1
            {'x': MARGIN_LEFT, 'y': A4_HEIGHT - MARGIN_TOP - 230, 'w': USABLE_WIDTH/3 - 7, 'h': 210},
            {'x': MARGIN_LEFT + USABLE_WIDTH/3, 'y': A4_HEIGHT - MARGIN_TOP - 230, 'w': USABLE_WIDTH/3 - 7, 'h': 210},
            {'x': MARGIN_LEFT + 2*USABLE_WIDTH/3, 'y': A4_HEIGHT - MARGIN_TOP - 230, 'w': USABLE_WIDTH/3 - 7, 'h': 210},
            # Row 2
            {'x': MARGIN_LEFT, 'y': A4_HEIGHT - MARGIN_TOP - 460, 'w': USABLE_WIDTH/3 - 7, 'h': 210},
            {'x': MARGIN_LEFT + USABLE_WIDTH/3, 'y': A4_HEIGHT - MARGIN_TOP - 460, 'w': USABLE_WIDTH/3 - 7, 'h': 210},
            {'x': MARGIN_LEFT + 2*USABLE_WIDTH/3, 'y': A4_HEIGHT - MARGIN_TOP - 460, 'w': USABLE_WIDTH/3 - 7, 'h': 210},
            # Row 3
            {'x': MARGIN_LEFT, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/3 - 7, 'h': 210},
            {'x': MARGIN_LEFT + USABLE_WIDTH/3, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/3 - 7, 'h': 210},
            {'x': MARGIN_LEFT + 2*USABLE_WIDTH/3, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/3 - 7, 'h': 210},
        ]
    },
    '1L4S': {
        'id': '1L4S',
        'name': '1 Large + 4 Small',
        'description': '1 large image on left, 4 small on right',
        'cells': 5,
        'grid': [
            # Large image on left
            {'x': MARGIN_LEFT, 'y': A4_HEIGHT - MARGIN_TOP - 500, 'w': USABLE_WIDTH * 0.55, 'h': 480},
            # 4 small images on right (2x2)
            {'x': MARGIN_LEFT + USABLE_WIDTH * 0.58, 'y': A4_HEIGHT - MARGIN_TOP - 240, 'w': USABLE_WIDTH * 0.4, 'h': 220},
            {'x': MARGIN_LEFT + USABLE_WIDTH * 0.58, 'y': A4_HEIGHT - MARGIN_TOP - 480, 'w': USABLE_WIDTH * 0.4, 'h': 220},
            # Bottom row
            {'x': MARGIN_LEFT, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/2 - 5, 'h': 200},
            {'x': MARGIN_LEFT + USABLE_WIDTH/2 + 5, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/2 - 5, 'h': 200},
        ]
    },
    '2L2S': {
        'id': '2L2S',
        'name': '2 Large + 2 Small',
        'description': '2 large images on top, 2 small below',
        'cells': 4,
        'grid': [
            # 2 large on top
            {'x': MARGIN_LEFT, 'y': A4_HEIGHT - MARGIN_TOP - 400, 'w': USABLE_WIDTH/2 - 5, 'h': 380},
            {'x': MARGIN_LEFT + USABLE_WIDTH/2 + 5, 'y': A4_HEIGHT - MARGIN_TOP - 400, 'w': USABLE_WIDTH/2 - 5, 'h': 380},
            # 2 small on bottom
            {'x': MARGIN_LEFT, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/2 - 5, 'h': 280},
            {'x': MARGIN_LEFT + USABLE_WIDTH/2 + 5, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/2 - 5, 'h': 280},
        ]
    },
    '3H': {
        'id': '3H',
        'name': '3 Horizontal',
        'description': '3 images stacked vertically',
        'cells': 3,
        'grid': [
            {'x': MARGIN_LEFT + USABLE_WIDTH * 0.1, 'y': A4_HEIGHT - MARGIN_TOP - 230, 'w': USABLE_WIDTH * 0.8, 'h': 210},
            {'x': MARGIN_LEFT + USABLE_WIDTH * 0.1, 'y': A4_HEIGHT - MARGIN_TOP - 460, 'w': USABLE_WIDTH * 0.8, 'h': 210},
            {'x': MARGIN_LEFT + USABLE_WIDTH * 0.1, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH * 0.8, 'h': 210},
        ]
    },
    '1XL': {
        'id': '1XL',
        'name': 'Single Large',
        'description': '1 full-page image',
        'cells': 1,
        'grid': [
            {'x': MARGIN_LEFT + 20, 'y': MARGIN_BOTTOM + 40, 'w': USABLE_WIDTH - 40, 'h': USABLE_HEIGHT - 60},
        ]
    },
    'MIX6': {
        'id': 'MIX6',
        'name': 'Mixed 6',
        'description': 'Asymmetric 6-image layout',
        'cells': 6,
        'grid': [
            # Top row: 1 large + 1 medium
            {'x': MARGIN_LEFT, 'y': A4_HEIGHT - MARGIN_TOP - 320, 'w': USABLE_WIDTH * 0.6, 'h': 300},
            {'x': MARGIN_LEFT + USABLE_WIDTH * 0.63, 'y': A4_HEIGHT - MARGIN_TOP - 320, 'w': USABLE_WIDTH * 0.35, 'h': 300},
            # Middle row: 2 equal
            {'x': MARGIN_LEFT, 'y': A4_HEIGHT - MARGIN_TOP - 540, 'w': USABLE_WIDTH/2 - 5, 'h': 200},
            {'x': MARGIN_LEFT + USABLE_WIDTH/2 + 5, 'y': A4_HEIGHT - MARGIN_TOP - 540, 'w': USABLE_WIDTH/2 - 5, 'h': 200},
            # Bottom row: 2 equal
            {'x': MARGIN_LEFT, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/2 - 5, 'h': 200},
            {'x': MARGIN_LEFT + USABLE_WIDTH/2 + 5, 'y': MARGIN_BOTTOM + 20, 'w': USABLE_WIDTH/2 - 5, 'h': 200},
        ]
    },
}


def get_layouts() -> List[Dict[str, Any]]:
    """Get list of available layouts for frontend."""
    return [
        {
            'id': layout['id'],
            'name': layout['name'],
            'description': layout['description'],
            'cells': layout['cells']
        }
        for layout in PREDEFINED_LAYOUTS.values()
    ]


def get_layout_by_id(layout_id: str) -> Optional[Dict[str, Any]]:
    """Get a specific layout by ID."""
    return PREDEFINED_LAYOUTS.get(layout_id)


class PlateGenerator:
    """Generates PDF plates with pottery images."""

    def __init__(self, output_dir: str = 'exports/plates'):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.temp_files = []

    def generate_plates(
        self,
        items: List[Dict[str, Any]],
        layout_id: str,
        period: str,
        group_by_us: bool = False,
        caption_format: str = 'inv',  # 'inv' or 'inv_us'
        start_plate_number: int = 1
    ) -> Tuple[str, List[Dict[str, Any]]]:
        """
        Generate PDF plates from classified items.

        Args:
            items: List of items with id, us, thumbnail, image_path
            layout_id: ID of the layout to use
            period: Period name for header
            group_by_us: Whether to group items by US
            caption_format: 'inv' for just ID, 'inv_us' for ID + US
            start_plate_number: Starting plate number

        Returns:
            Tuple of (PDF filename, plate assignments for Excel report)
        """
        layout = get_layout_by_id(layout_id)
        if not layout:
            raise ValueError(f"Unknown layout: {layout_id}")

        cells_per_plate = layout['cells']

        # Split items into plates, optionally grouped by US
        plates = []  # List of (us_label, items) tuples

        if group_by_us:
            grouped = {}
            for item in items:
                us = item.get('us', 'Unknown')
                if us not in grouped:
                    grouped[us] = []
                grouped[us].append(item)
            # Create plates per US group
            for us in sorted(grouped.keys()):
                us_items = grouped[us]
                for i in range(0, len(us_items), cells_per_plate):
                    plates.append((f"US {us}", us_items[i:i + cells_per_plate]))
        else:
            # No grouping
            for i in range(0, len(items), cells_per_plate):
                plates.append((None, items[i:i + cells_per_plate]))

        if not plates:
            plates = [(None, [])]  # At least one empty plate

        # Generate PDF
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'plates_{period.replace(" ", "_")}_{timestamp}.pdf'
        output_path = self.output_dir / filename

        c = canvas.Canvas(str(output_path), pagesize=A4)

        plate_assignments = []

        for plate_idx, (us_label, plate_items) in enumerate(plates):
            plate_number = start_plate_number + plate_idx

            # Draw header with optional US
            header_text = period
            if us_label:
                header_text = f"{period} - {us_label}"
            self._draw_header(c, plate_number, header_text)

            # Draw items in cells
            for cell_idx, item in enumerate(plate_items):
                if cell_idx >= len(layout['grid']):
                    break

                cell = layout['grid'][cell_idx]
                caption = self._format_caption(item, caption_format)

                self._draw_item_in_cell(c, item, cell, caption)

                # Record assignment
                plate_assignments.append({
                    'plate_number': plate_number,
                    'position': cell_idx + 1,
                    'id': item.get('id', ''),
                    'us': item.get('us', ''),
                    'period': period,
                    'layout': layout_id
                })

            # New page if not last plate
            if plate_idx < len(plates) - 1:
                c.showPage()

        c.save()

        # Cleanup temp files
        self._cleanup_temp_files()

        return filename, plate_assignments

    def _draw_header(self, c: canvas.Canvas, plate_number: int, period: str):
        """Draw plate header with number and period."""
        c.setFont('Helvetica-Bold', 14)
        c.drawCentredString(A4_WIDTH / 2, A4_HEIGHT - 15 * mm, f'PLATE {plate_number}')
        c.setFont('Helvetica', 11)
        c.drawCentredString(A4_WIDTH / 2, A4_HEIGHT - 22 * mm, period)

    def _format_caption(self, item: Dict[str, Any], caption_format: str) -> str:
        """Format caption based on format type."""
        item_id = item.get('id', 'N/A')
        us = item.get('us', '')
        sito = item.get('sito', '')

        if caption_format in ('inv_us', 'id_us') and us:
            return f"{item_id} - US {us}"
        elif caption_format in ('inv_us_sito', 'id_us_sito'):
            parts = [str(item_id)]
            if us:
                parts.append(f"US {us}")
            if sito:
                parts.append(sito)
            return " - ".join(parts)
        return str(item_id)

    def _draw_item_in_cell(
        self,
        c: canvas.Canvas,
        item: Dict[str, Any],
        cell: Dict[str, float],
        caption: str
    ):
        """Draw an item (image + caption) in a cell."""
        x, y, w, h = cell['x'], cell['y'], cell['w'], cell['h']

        # Reserve space for caption
        img_height = h - CAPTION_HEIGHT - 5

        # Try to load and draw image
        image_drawn = False

        # Try image_path first (high resolution)
        if item.get('image_path') and os.path.exists(item['image_path']):
            try:
                self._draw_image_scaled(c, item['image_path'], x, y + CAPTION_HEIGHT + 5, w, img_height)
                image_drawn = True
            except Exception as e:
                print(f"Error drawing image from path: {e}")

        # Fallback to thumbnail (base64)
        if not image_drawn and item.get('thumbnail'):
            try:
                temp_path = self._base64_to_temp_file(item['thumbnail'])
                if temp_path:
                    self._draw_image_scaled(c, temp_path, x, y + CAPTION_HEIGHT + 5, w, img_height)
                    image_drawn = True
            except Exception as e:
                print(f"Error drawing thumbnail: {e}")

        # Draw placeholder if no image
        if not image_drawn:
            c.setStrokeColorRGB(0.7, 0.7, 0.7)
            c.setFillColorRGB(0.95, 0.95, 0.95)
            c.rect(x, y + CAPTION_HEIGHT + 5, w, img_height, fill=1, stroke=1)
            c.setFillColorRGB(0.5, 0.5, 0.5)
            c.setFont('Helvetica', 10)
            c.drawCentredString(x + w/2, y + CAPTION_HEIGHT + 5 + img_height/2, 'No Image')

        # Draw caption
        c.setFillColorRGB(0, 0, 0)
        c.setFont(CAPTION_FONT, CAPTION_SIZE)
        c.drawCentredString(x + w/2, y + 3, caption)

    def _draw_image_scaled(
        self,
        c: canvas.Canvas,
        image_path: str,
        x: float,
        y: float,
        max_width: float,
        max_height: float
    ):
        """Draw an image scaled to fit within max dimensions, centered."""
        img = Image.open(image_path)
        img_width, img_height = img.size

        # Calculate scale to fit
        scale_w = max_width / img_width
        scale_h = max_height / img_height
        scale = min(scale_w, scale_h)

        new_width = img_width * scale
        new_height = img_height * scale

        # Center in cell
        offset_x = (max_width - new_width) / 2
        offset_y = (max_height - new_height) / 2

        c.drawImage(
            image_path,
            x + offset_x,
            y + offset_y,
            width=new_width,
            height=new_height,
            preserveAspectRatio=True
        )

    def _base64_to_temp_file(self, base64_data: str) -> Optional[str]:
        """Convert base64 image data to temporary file."""
        try:
            # Remove data URI prefix if present
            if ',' in base64_data:
                base64_data = base64_data.split(',')[1]

            image_data = base64.b64decode(base64_data)

            # Create temp file
            fd, temp_path = tempfile.mkstemp(suffix='.jpg')
            os.close(fd)

            with open(temp_path, 'wb') as f:
                f.write(image_data)

            self.temp_files.append(temp_path)
            return temp_path

        except Exception as e:
            print(f"Error converting base64 to temp file: {e}")
            return None

    def _cleanup_temp_files(self):
        """Remove temporary files."""
        for temp_path in self.temp_files:
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except Exception:
                pass
        self.temp_files = []

    def generate_preview(
        self,
        items: List[Dict[str, Any]],
        layout_id: str,
        caption_format: str = 'inv'
    ) -> str:
        """
        Generate a preview image (PNG) of a single plate.

        Returns:
            Base64-encoded PNG image
        """
        layout = get_layout_by_id(layout_id)
        if not layout:
            raise ValueError(f"Unknown layout: {layout_id}")

        # Create preview using PIL
        # Scale factor for preview (A4 at 72 DPI)
        width = int(A4_WIDTH)
        height = int(A4_HEIGHT)

        preview = Image.new('RGB', (width, height), 'white')

        # Draw cells with items
        from PIL import ImageDraw, ImageFont
        draw = ImageDraw.Draw(preview)

        for cell_idx, cell in enumerate(layout['grid']):
            x, y_bottom, w, h = int(cell['x']), int(cell['y']), int(cell['w']), int(cell['h'])
            # Convert from bottom-left to top-left coordinate system
            y_top = height - y_bottom - int(h)

            # Draw cell border
            draw.rectangle([x, y_top, x + w, y_top + int(h)], outline='#cccccc', width=1)

            # Draw item if available
            if cell_idx < len(items):
                item = items[cell_idx]

                # Try to draw thumbnail
                if item.get('thumbnail'):
                    try:
                        thumb_data = item['thumbnail']
                        if ',' in thumb_data:
                            thumb_data = thumb_data.split(',')[1]

                        thumb_bytes = base64.b64decode(thumb_data)
                        thumb_img = Image.open(io.BytesIO(thumb_bytes))

                        # Scale to fit cell (with margin for caption)
                        cell_w = w - 10
                        cell_h = int(h) - CAPTION_HEIGHT - 10

                        thumb_img.thumbnail((cell_w, cell_h), Image.Resampling.LANCZOS)

                        # Center in cell
                        thumb_w, thumb_h = thumb_img.size
                        paste_x = x + (w - thumb_w) // 2
                        paste_y = y_top + (int(h) - CAPTION_HEIGHT - thumb_h) // 2

                        # Handle RGBA
                        if thumb_img.mode == 'RGBA':
                            preview.paste(thumb_img, (paste_x, paste_y), thumb_img)
                        else:
                            preview.paste(thumb_img, (paste_x, paste_y))

                    except Exception as e:
                        print(f"Error drawing preview thumbnail: {e}")

                # Draw caption
                caption = self._format_caption(item, caption_format)
                caption_y = y_top + int(h) - CAPTION_HEIGHT

                # Simple text (no custom font in preview)
                try:
                    font = ImageFont.truetype("/System/Library/Fonts/Helvetica.ttc", 9)
                except:
                    font = ImageFont.load_default()

                bbox = draw.textbbox((0, 0), caption, font=font)
                text_width = bbox[2] - bbox[0]
                text_x = x + (w - text_width) // 2
                draw.text((text_x, caption_y), caption, fill='black', font=font)

        # Draw header area
        draw.text((width // 2 - 30, 15), "PLATE X", fill='#666666')

        # Convert to base64
        buffer = io.BytesIO()
        preview.save(buffer, format='PNG')
        buffer.seek(0)

        return 'data:image/png;base64,' + base64.b64encode(buffer.read()).decode()


def generate_excel_report(
    plate_assignments: List[Dict[str, Any]],
    output_dir: str = 'exports/plates'
) -> str:
    """
    Generate Excel report with plate assignments.

    Args:
        plate_assignments: List of assignment dicts from generate_plates
        output_dir: Output directory

    Returns:
        Filename of generated Excel file
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'plate_report_{timestamp}.xlsx'
    filepath = output_path / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Plate Assignments'

    # Headers
    headers = ['Plate #', 'Position', 'Inventory ID', 'US', 'Period', 'Layout']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for row_idx, assignment in enumerate(plate_assignments, 2):
        ws.cell(row=row_idx, column=1, value=assignment['plate_number'])
        ws.cell(row=row_idx, column=2, value=assignment['position'])
        ws.cell(row=row_idx, column=3, value=assignment['id'])
        ws.cell(row=row_idx, column=4, value=assignment.get('us', ''))
        ws.cell(row=row_idx, column=5, value=assignment['period'])
        ws.cell(row=row_idx, column=6, value=assignment['layout'])

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12

    # Freeze header
    ws.freeze_panes = 'A2'

    wb.save(filepath)

    return filename
