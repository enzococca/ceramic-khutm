#!/usr/bin/env python3
"""
Ceramica KhUTM - Web Application for Pottery Classification
Real-time monitoring with WebSocket and Plotly charts
"""

from flask import Flask, render_template, jsonify, request, send_file
from flask_socketio import SocketIO, emit
import psycopg2
import requests
import base64
import os
import json
import threading
import time
from pathlib import Path
from PIL import Image
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
import traceback

from plate_generator import PlateGenerator, get_layouts, get_layout_by_id, generate_excel_report

app = Flask(__name__)
app.config['SECRET_KEY'] = 'ceramica-khutm-secret-2024'
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='eventlet')

# Configuration
DB_CONFIG = {
    'host': 'localhost',
    'port': 5433,
    'database': 'khutm2',
    'user': 'postgres',
    'password': 'postgres'
}

PHOTOLOG_DIR = Path("/Volumes/extesione4T/KTM2025/photolog/original")
API_URL = "https://pottery-comparison-oman.up.railway.app/api/ml/similar"
EXPORT_DIR = Path(__file__).parent / "exports"
THUMBNAIL_SIZE = (100, 100)

# Global state
classification_state = {
    'running': False,
    'paused': False,
    'total': 0,
    'processed': 0,
    'errors': 0,
    'current_item': None,
    'results': [],
    'statistics': {
        'periods': {},
        'decorations': {},
        'confidences': [],
        'sites': {}
    },
    'start_time': None,
    'last_export': None
}


def get_db_connection():
    """Get database connection."""
    return psycopg2.connect(**DB_CONFIG)


def get_decorated_pottery():
    """Get all decorated pottery with their associated images."""
    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT DISTINCT
            p.id_rep,
            p.sito,
            p.area,
            p.us,
            p.box,
            p.form,
            p.specific_form,
            p.ware,
            p.fabric,
            p.exdeco,
            p.intdeco,
            p.decoration_type,
            p.decoration_motif,
            p.decoration_position,
            p.descrip_ex_deco,
            p.descrip_in_deco,
            p.datazione,
            p.note,
            mte.id_media,
            m.filename
        FROM pottery_table p
        JOIN media_to_entity_table mte ON mte.id_entity = p.id_rep
            AND mte.entity_type = 'CERAMICA'
        JOIN media_table m ON m.id_media = mte.id_media
        WHERE p.exdeco = 'Yes'
           OR p.intdeco = 'Yes'
           OR (p.decoration_type IS NOT NULL AND p.decoration_type != '' AND p.decoration_type != 'Slipped')
        ORDER BY p.id_rep, mte.id_media
    """

    cur.execute(query)
    rows = cur.fetchall()

    columns = [
        'id_rep', 'sito', 'area', 'us', 'box', 'form', 'specific_form',
        'ware', 'fabric', 'exdeco', 'intdeco', 'decoration_type',
        'decoration_motif', 'decoration_position', 'descrip_ex_deco',
        'descrip_in_deco', 'datazione', 'note', 'id_media', 'filename'
    ]

    results = []
    for row in rows:
        results.append(dict(zip(columns, row)))

    conn.close()
    return results


def find_image_file(id_media, filename):
    """Find the image file in the photolog directory."""
    pattern = f"{id_media}_{filename}.png"
    file_path = PHOTOLOG_DIR / pattern

    if file_path.exists():
        return file_path

    for ext in ['.png', '.PNG', '.jpg', '.JPG', '.jpeg', '.JPEG']:
        pattern = f"{id_media}_{filename}{ext}"
        file_path = PHOTOLOG_DIR / pattern
        if file_path.exists():
            return file_path

    for f in PHOTOLOG_DIR.glob(f"{id_media}_*"):
        if not f.name.startswith('._'):
            return f

    return None


def image_to_base64(image_path, max_size=800):
    """Convert image file to base64 data URI with optional resizing."""
    with Image.open(image_path) as img:
        # Resize if too large
        if max(img.size) > max_size:
            img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)

        # Convert to RGB if necessary
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')

        buffer = io.BytesIO()
        img.save(buffer, format='JPEG', quality=85)
        b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

    return f"data:image/jpeg;base64,{b64}"


def create_thumbnail_base64(image_path, size=(150, 150)):
    """Create a base64 thumbnail from the image."""
    try:
        with Image.open(image_path) as img:
            img.thumbnail(size, Image.Resampling.LANCZOS)

            if img.mode in ('RGBA', 'P'):
                img = img.convert('RGB')

            buffer = io.BytesIO()
            img.save(buffer, format='JPEG', quality=80)
            b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

        return f"data:image/jpeg;base64,{b64}"
    except Exception as e:
        return None


def classify_image(image_path, top_k=5):
    """Classify an image using the ML similarity API."""
    try:
        image_data = image_to_base64(image_path)

        response = requests.post(
            API_URL,
            json={'image': image_data, 'top_k': top_k},
            timeout=120
        )

        if response.status_code == 200:
            result = response.json()
            if result.get('success') and result.get('analysis'):
                return {
                    'success': True,
                    'period_suggestion': result['analysis'].get('period_suggestion', ''),
                    'period_confidence': result['analysis'].get('period_confidence', 0),
                    'decoration_suggestion': result['analysis'].get('decoration_suggestion', ''),
                    'sites': result['analysis'].get('sites', []),
                    'top_match': result['similar_items'][0] if result.get('similar_items') else None,
                    'similar_items': result.get('similar_items', [])[:5],
                    'analysis_text': result['analysis'].get('text', '')
                }

        return {'success': False, 'error': f"API error: {response.status_code}"}

    except Exception as e:
        return {'success': False, 'error': str(e)}


def update_statistics(classification):
    """Update running statistics."""
    global classification_state

    if classification.get('success'):
        period = classification.get('period_suggestion', 'Unknown')
        decoration = classification.get('decoration_suggestion', 'Unknown')
        confidence = classification.get('period_confidence', 0)

        # Period distribution
        classification_state['statistics']['periods'][period] = \
            classification_state['statistics']['periods'].get(period, 0) + 1

        # Decoration distribution
        classification_state['statistics']['decorations'][decoration] = \
            classification_state['statistics']['decorations'].get(decoration, 0) + 1

        # Confidence values
        classification_state['statistics']['confidences'].append(confidence)

        # Sites distribution
        for site in classification.get('sites', []):
            classification_state['statistics']['sites'][site] = \
                classification_state['statistics']['sites'].get(site, 0) + 1


def run_classification(limit=None):
    """Run the classification process."""
    global classification_state

    try:
        # Reset state
        classification_state['running'] = True
        classification_state['paused'] = False
        classification_state['processed'] = 0
        classification_state['errors'] = 0
        classification_state['results'] = []
        classification_state['statistics'] = {
            'periods': {},
            'decorations': {},
            'confidences': [],
            'sites': {}
        }
        classification_state['start_time'] = datetime.now()

        socketio.emit('status', {'status': 'starting', 'message': 'Caricamento ceramiche dal database...'})

        # Get pottery items
        pottery_items = get_decorated_pottery()

        # Get unique pottery (group by id_rep)
        unique_pottery = {}
        for item in pottery_items:
            id_rep = item['id_rep']
            if id_rep not in unique_pottery:
                unique_pottery[id_rep] = item

        if limit:
            unique_pottery = dict(list(unique_pottery.items())[:limit])

        classification_state['total'] = len(unique_pottery)

        socketio.emit('status', {
            'status': 'running',
            'message': f'Trovate {len(unique_pottery)} ceramiche decorate',
            'total': len(unique_pottery)
        })

        # Process each item
        for idx, (id_rep, item) in enumerate(unique_pottery.items()):
            # Check if paused or stopped
            while classification_state['paused'] and classification_state['running']:
                socketio.sleep(0.5)

            if not classification_state['running']:
                break

            # Find image
            image_path = find_image_file(item['id_media'], item['filename'])

            result_item = {
                'id_rep': id_rep,
                'sito': item.get('sito'),
                'area': item.get('area'),
                'us': item.get('us'),
                'box': item.get('box'),
                'form': item.get('form'),
                'specific_form': item.get('specific_form'),
                'ware': item.get('ware'),
                'fabric': item.get('fabric'),
                'exdeco': item.get('exdeco'),
                'intdeco': item.get('intdeco'),
                'decoration_type': item.get('decoration_type'),
                'decoration_motif': item.get('decoration_motif'),
                'filename': item.get('filename'),
                'id_media': item.get('id_media')
            }

            if image_path:
                # Create thumbnail for display
                thumbnail = create_thumbnail_base64(image_path)
                result_item['thumbnail'] = thumbnail

                # Classify
                classification = classify_image(image_path)
                result_item['classification'] = classification

                if classification.get('success'):
                    update_statistics(classification)
                    result_item['status'] = 'success'
                else:
                    classification_state['errors'] += 1
                    result_item['status'] = 'error'
                    result_item['error'] = classification.get('error', 'Unknown error')
            else:
                classification_state['errors'] += 1
                result_item['status'] = 'error'
                result_item['error'] = 'Image not found'
                result_item['classification'] = {'success': False, 'error': 'Image not found'}

            classification_state['results'].append(result_item)
            classification_state['processed'] += 1
            classification_state['current_item'] = result_item

            # Emit progress update
            socketio.emit('progress', {
                'processed': classification_state['processed'],
                'total': classification_state['total'],
                'errors': classification_state['errors'],
                'current_item': result_item,
                'statistics': classification_state['statistics']
            })

            # Small delay to not overload
            socketio.sleep(0.1)

        # Finished
        classification_state['running'] = False
        elapsed = (datetime.now() - classification_state['start_time']).total_seconds()

        socketio.emit('status', {
            'status': 'completed',
            'message': f'Classificazione completata in {elapsed:.1f} secondi',
            'processed': classification_state['processed'],
            'errors': classification_state['errors']
        })

    except Exception as e:
        classification_state['running'] = False
        socketio.emit('status', {
            'status': 'error',
            'message': f'Errore: {str(e)}'
        })
        traceback.print_exc()


def export_to_excel():
    """Export results to Excel with thumbnails."""
    global classification_state

    if not classification_state['results']:
        return None

    EXPORT_DIR.mkdir(exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pottery Classification"

    headers = [
        'Thumbnail', 'ID Rep', 'Sito', 'Area', 'US', 'Box', 'Form', 'Specific Form',
        'Ware', 'Fabric', 'Ext Deco', 'Int Deco', 'Decoration Type',
        'Decoration Motif', 'Image Filename', 'ML Period', 'ML Confidence',
        'ML Decoration', 'ML Sites', 'Top Match ID', 'Top Match Similarity',
        'Status'
    ]

    # Write headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths
    ws.column_dimensions['A'].width = 12  # Thumbnail
    ws.column_dimensions['B'].width = 10  # ID Rep
    ws.column_dimensions['C'].width = 12  # Sito
    ws.column_dimensions['O'].width = 18  # Filename
    ws.column_dimensions['P'].width = 15  # ML Period
    ws.column_dimensions['S'].width = 25  # ML Sites

    # Temp files to cleanup
    temp_files = []

    # Write data
    for row_num, item in enumerate(classification_state['results'], 2):
        # Set row height for thumbnail
        ws.row_dimensions[row_num].height = 60

        classification = item.get('classification', {})
        top_match = classification.get('top_match', {}) or {}

        # Try to add thumbnail
        thumbnail_data = item.get('thumbnail')
        if thumbnail_data and thumbnail_data.startswith('data:image'):
            try:
                # Decode base64 thumbnail
                b64_data = thumbnail_data.split(',')[1]
                img_bytes = base64.b64decode(b64_data)

                # Save to temp file
                temp_path = tempfile.mktemp(suffix='.jpg')
                with open(temp_path, 'wb') as f:
                    f.write(img_bytes)
                temp_files.append(temp_path)

                # Add to Excel
                img = XLImage(temp_path)
                img.width = 60
                img.height = 60
                ws.add_image(img, f'A{row_num}')
            except Exception as e:
                print(f"Error adding thumbnail: {e}")

        row_data = [
            '',  # Thumbnail placeholder
            item.get('id_rep'),
            item.get('sito'),
            item.get('area'),
            item.get('us'),
            item.get('box'),
            item.get('form'),
            item.get('specific_form'),
            item.get('ware'),
            item.get('fabric'),
            item.get('exdeco'),
            item.get('intdeco'),
            item.get('decoration_type'),
            item.get('decoration_motif'),
            item.get('filename'),
            classification.get('period_suggestion', ''),
            classification.get('period_confidence', ''),
            classification.get('decoration_suggestion', ''),
            ', '.join(classification.get('sites', [])),
            top_match.get('id', ''),
            f"{top_match.get('similarity', 0):.1f}%" if top_match.get('similarity') else '',
            item.get('status', '')
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = Alignment(vertical='center')

    # Freeze header
    ws.freeze_panes = 'A2'

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = EXPORT_DIR / f"khutm_classification_{timestamp}.xlsx"
    wb.save(output_path)

    # Cleanup temp files
    for temp_file in temp_files:
        try:
            os.remove(temp_file)
        except:
            pass

    classification_state['last_export'] = str(output_path)
    return str(output_path)


# Routes
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/status')
def get_status():
    return jsonify({
        'running': classification_state['running'],
        'paused': classification_state['paused'],
        'total': classification_state['total'],
        'processed': classification_state['processed'],
        'errors': classification_state['errors'],
        'statistics': classification_state['statistics']
    })


@app.route('/api/results')
def get_results():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    start = (page - 1) * per_page
    end = start + per_page

    return jsonify({
        'results': classification_state['results'][start:end],
        'total': len(classification_state['results']),
        'page': page,
        'per_page': per_page
    })


@app.route('/api/export')
def export_results():
    path = export_to_excel()
    if path:
        return jsonify({'success': True, 'path': path})
    return jsonify({'success': False, 'error': 'No results to export'})


@app.route('/api/download')
def download_export():
    if classification_state['last_export'] and os.path.exists(classification_state['last_export']):
        return send_file(classification_state['last_export'], as_attachment=True)
    return jsonify({'error': 'No export available'}), 404


@app.route('/api/db-stats')
def db_stats():
    """Get database statistics."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # Total pottery
        cur.execute("SELECT COUNT(*) FROM pottery_table")
        total = cur.fetchone()[0]

        # Decorated
        cur.execute("""
            SELECT COUNT(*) FROM pottery_table
            WHERE exdeco = 'Yes' OR intdeco = 'Yes'
               OR (decoration_type IS NOT NULL AND decoration_type != '' AND decoration_type != 'Slipped')
        """)
        decorated = cur.fetchone()[0]

        # With media
        cur.execute("""
            SELECT COUNT(DISTINCT p.id_rep)
            FROM pottery_table p
            JOIN media_to_entity_table mte ON mte.id_entity = p.id_rep AND mte.entity_type = 'CERAMICA'
        """)
        with_media = cur.fetchone()[0]

        # Decoration types
        cur.execute("""
            SELECT decoration_type, COUNT(*)
            FROM pottery_table
            WHERE decoration_type IS NOT NULL AND decoration_type != ''
            GROUP BY decoration_type
            ORDER BY COUNT(*) DESC
        """)
        decoration_types = dict(cur.fetchall())

        conn.close()

        return jsonify({
            'total': total,
            'decorated': decorated,
            'with_media': with_media,
            'decoration_types': decoration_types
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============== Plate Generator API ==============

PLATES_DIR = Path(__file__).parent / "exports" / "plates"
PLATES_DIR.mkdir(parents=True, exist_ok=True)


@app.route('/api/plates/periods')
def api_plates_periods():
    """Get available periods from classification results."""
    results = classification_state.get('results', [])
    periods = {}

    for result in results:
        if result.get('status') == 'success':
            classification = result.get('classification', {})
            period = classification.get('period_suggestion', 'Unknown')
            if period:
                periods[period] = periods.get(period, 0) + 1

    return jsonify({
        'success': True,
        'periods': [{'name': k, 'count': v} for k, v in sorted(periods.items())]
    })


@app.route('/api/plates/items')
def api_plates_items():
    """Get classified items filtered by period and/or US."""
    period = request.args.get('period')
    us_filter = request.args.get('us')

    results = classification_state.get('results', [])
    items = []

    for result in results:
        if result.get('status') != 'success':
            continue

        classification = result.get('classification', {})
        item_period = classification.get('period_suggestion', '')

        if period and item_period != period:
            continue

        item_us = result.get('us', '')
        if us_filter and str(item_us) != str(us_filter):
            continue

        items.append({
            'id': result.get('id_rep', 'N/A'),
            'us': item_us,
            'period': item_period,
            'thumbnail': result.get('thumbnail', ''),
            'image_path': '',
            'form': result.get('form', ''),
            'decoration': result.get('decoration_type', '')
        })

    return jsonify({
        'success': True,
        'items': items,
        'count': len(items)
    })


@app.route('/api/plates/layouts')
def api_plates_layouts():
    """Get available layout definitions."""
    return jsonify({
        'success': True,
        'layouts': get_layouts()
    })


@app.route('/api/plates/preview', methods=['POST'])
def api_plates_preview():
    """Generate preview image for a plate."""
    try:
        data = request.json
        item_ids = data.get('items', [])
        layout_id = data.get('layout_id', '2x2')
        caption_format = data.get('caption_format', 'inv')

        # Convert IDs to full item objects from classification results
        results = classification_state.get('results', [])
        items = []
        for result in results:
            if result.get('status') == 'success':
                item_id = result.get('id_rep', '')
                if str(item_id) in [str(i) for i in item_ids]:
                    classification = result.get('classification', {})
                    items.append({
                        'id': item_id,
                        'us': result.get('us', ''),
                        'sito': result.get('sito', ''),
                        'period': classification.get('period_suggestion', ''),
                        'thumbnail': result.get('thumbnail', ''),
                        'image_path': result.get('image_path', '')
                    })

        generator = PlateGenerator(str(PLATES_DIR))
        preview = generator.generate_preview(items, layout_id, caption_format)

        return jsonify({
            'success': True,
            'preview': preview
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        })


@app.route('/api/plates/generate', methods=['POST'])
def api_plates_generate():
    """Generate PDF plates."""
    try:
        data = request.json
        item_ids = data.get('items', [])
        layout_id = data.get('layout_id', '2x2')
        period = data.get('period', 'Unknown')
        group_by_us = data.get('group_by_us', False)
        caption_format = data.get('caption_format', 'inv')
        start_plate_number = data.get('start_plate_number', 1)

        # Convert IDs to full item objects from classification results
        results = classification_state.get('results', [])
        items = []
        for result in results:
            if result.get('status') == 'success':
                item_id = result.get('id_rep', '')
                if str(item_id) in [str(i) for i in item_ids]:
                    classification = result.get('classification', {})
                    items.append({
                        'id': item_id,
                        'us': result.get('us', ''),
                        'sito': result.get('sito', ''),
                        'period': classification.get('period_suggestion', ''),
                        'thumbnail': result.get('thumbnail', ''),
                        'image_path': result.get('image_path', '')
                    })

        generator = PlateGenerator(str(PLATES_DIR))
        filename, assignments = generator.generate_plates(
            items=items,
            layout_id=layout_id,
            period=period,
            group_by_us=group_by_us,
            caption_format=caption_format,
            start_plate_number=start_plate_number
        )

        return jsonify({
            'success': True,
            'filename': filename,
            'plates_count': len(set(a['plate_number'] for a in assignments)),
            'items_count': len(assignments)
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        })


@app.route('/api/plates/report', methods=['POST'])
def api_plates_report():
    """Generate Excel report for plate assignments."""
    try:
        data = request.json
        plates = data.get('plates', [])

        # Convert nested plates data to flat assignments format
        assignments = []
        for plate in plates:
            plate_number = plate.get('plate_number', 0)
            period = plate.get('period', '')
            us = plate.get('us', '')
            items = plate.get('items', [])

            for pos, item in enumerate(items, 1):
                assignments.append({
                    'plate_number': plate_number,
                    'position': pos,
                    'id': item.get('id', ''),
                    'us': item.get('us', us),
                    'period': period,
                    'layout': data.get('layout_id', '2x3')
                })

        filename = generate_excel_report(assignments, str(PLATES_DIR))

        return jsonify({
            'success': True,
            'filename': filename
        })

    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        })


@app.route('/api/plates/download/<filename>')
def api_plates_download(filename):
    """Download generated plate file."""
    filepath = PLATES_DIR / filename
    if filepath.exists():
        return send_file(filepath, as_attachment=True)
    return jsonify({'error': 'File not found'}), 404


# Socket events
@socketio.on('connect')
def handle_connect():
    emit('status', {
        'status': 'connected',
        'running': classification_state['running'],
        'processed': classification_state['processed'],
        'total': classification_state['total']
    })


@socketio.on('start')
def handle_start(data):
    if classification_state['running']:
        emit('status', {'status': 'error', 'message': 'Classificazione già in corso'})
        return

    limit = data.get('limit')
    socketio.start_background_task(run_classification, limit)


@socketio.on('pause')
def handle_pause():
    classification_state['paused'] = not classification_state['paused']
    emit('status', {
        'status': 'paused' if classification_state['paused'] else 'running',
        'message': 'In pausa' if classification_state['paused'] else 'Ripresa'
    })


@socketio.on('stop')
def handle_stop():
    classification_state['running'] = False
    emit('status', {'status': 'stopped', 'message': 'Classificazione interrotta'})


if __name__ == '__main__':
    print("=" * 60)
    print("Ceramica KhUTM - Classification Web App")
    print("=" * 60)
    print(f"Server running at: http://localhost:5001")
    print(f"Photolog directory: {PHOTOLOG_DIR}")
    print(f"ML API: {API_URL}")
    print("=" * 60)

    socketio.run(app, host='0.0.0.0', port=5001, debug=True)
